from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import os
import pandas as pd
from datetime import timedelta, date, time, datetime
os.chdir("C:\\Users\\Ian\\OneDrive - University of Pittsburgh\\Career\\SCS Employment\\Timecard automation")

#TODO: Ask for file name, make output file the correct name (based on date)
#load login times and the schedule excel docs
wb1 = load_workbook("schedule.xlsx",read_only=False)
report = wb1.active

keys = pd.read_excel(io="PGH - Timecard Check - Weekly.xlsx")
keys = keys.drop(columns=["idnt","agid","Duration", "Name"]) #remove useless cols
# keys = keys.sort_values(['User', 'Logon'])


# Traverse through report sheet, verifying schedule with keys sheet
report_row = 1

# Loops for each employee
while report.cell(row=report_row, column=1).value != None: #TODO: contains full name instead of ==
    #get username and full name from curr row and col
    raw_cell_name = report.cell(row=report_row, column=1).value
    if (raw_cell_name == None):
        break
    allParts = raw_cell_name.split(sep="|")
    
    full_name = allParts[0][:-1]
    username = allParts[1][1:]
    print("Full name: " + full_name + "\nusername: " + username)
    allName = full_name.split(sep=',')
    firstName = allName[1][1:]
    lastName = allName[0]

    user_logins = []
    user_logouts = []
    for row in range(len(keys.index)): #Get all logins by employee
        if keys.loc[row][0].casefold() == username.casefold() or (firstName in keys.loc[row][0] and lastName in keys.loc[row][0]):
            user_logins.append(keys.loc[row][1])
            user_logouts.append(keys.loc[row][2])
    user_logins.sort()
    user_logouts.sort()
    #print(user_logins + user_logouts)
    
    
    #Loops for each shift for an employee
    report_row += 1
    currVal = report.cell(row=report_row, column= 1).value
    while currVal != None: # None means end of this employees shifts, 
        #Look at single shift (or connected shifts)
        num_connected = 0
        date = report.cell(row=report_row, column=1).value
        start = datetime.combine(date, report.cell(row=report_row, column=2).value)
        end = datetime.combine(date, report.cell(row=report_row, column=3).value)
        # While end time is equal to start time of next shift and dates are equal, set end time to next shift end
        while (report.cell(row=report_row+1, column=1).value != None 
                and end >= datetime.combine(report.cell(row=report_row+1, column=1).value, report.cell(row=report_row+1, column=2).value) - timedelta(hours=1)): # Combines shifts within 1 hours of eachother
            report_row += 1
            num_connected += 1
            end = datetime.combine(date, report.cell(row=report_row, column=3).value)
        
        # check keys for earliest logon that matches
        noLogins = False
        if len(user_logins) < 1: #no logins found for the week so trigger no show 
            noLogins = True
        else:
            closest_start_time = user_logins[0]
            for time in user_logins:
                time_diff = start - time
                if abs(time_diff) < abs(start - closest_start_time):
                    closest_start_time = time

            # Check for closest logoff
            closest_end_time = user_logouts[0]
            for time in user_logouts:
                time_diff = end - time
                if abs(time_diff) < abs(end - closest_end_time):
                    closest_end_time = time

        print("Checking " + str(num_connected+1) + " shift(s): " + str(start) + " to " + str(end))
        #Edit sheet based on results
        
        currRow = report_row - num_connected
        currStart = datetime.combine(report.cell(row=currRow, column=1).value, report.cell(row=currRow, column=2).value)
        currEnd = datetime.combine(report.cell(row=currRow, column=1).value, report.cell(row=currRow, column=3).value)
        while currRow != report_row + 1:
            if noLogins or closest_start_time >= currEnd or closest_end_time <= currStart: #If no show 
                for col in range(1,7):
                    report.cell(row=currRow, column=col).fill = PatternFill(fgColor="FF0000", fill_type="solid") #fill red
            elif closest_start_time > currStart + timedelta(minutes=6) and closest_start_time < currEnd: #If late
                for col in range(1,7): 
                        report.cell(row=currRow, column=col).fill = PatternFill(fgColor="FFFF00", fill_type="solid") #fill yellow
                #print login time
                zero_str = ""
                hour = closest_start_time.hour
                if (hour > 12):
                    hour -= 12
                if (closest_start_time.minute < 10):
                    zero_str = "0"
                start_time_str = str(hour) + ":" + zero_str + str(closest_start_time.minute)
                report.cell(row=currRow,column=7).value = start_time_str
            elif closest_end_time < currEnd - timedelta(minutes=6) and closest_end_time > currStart: #If leave early
                for col in range(1,7): 
                    report.cell(row=currRow, column=col).fill = PatternFill(fgColor="FFFF00", fill_type="solid") #Fill yellow
                #print logout time
                zero_str = ""
                hour = closest_end_time.hour
                if (closest_end_time.minute < 10):
                    zero_str = "0"
                if hour > 12:
                    hour -= 12
                end_time_str = str(hour) + ":" + zero_str + str(closest_end_time.minute)
                report.cell(row=currRow, column=8).value = end_time_str
            else: # Valid Shift
                for col in range(1,7):
                    report.cell(row=currRow, column=col).fill = PatternFill(fgColor="92D050", fill_type="solid") #fill green

            currRow += 1
            if report.cell(row=currRow, column=1).value == None:
                break
            currStart = datetime.combine(report.cell(row=currRow, column=1).value, report.cell(row=currRow, column=2).value)
            currEnd = datetime.combine(report.cell(row=currRow, column=1).value, report.cell(row=currRow, column=3).value)
        print()
        
        report_row += 1
        currVal = report.cell(row=report_row, column= 1).value
    
    report_row += 1

print("Processed all Shifts, Report completed")
Workbook.save(self=wb1, filename="2024-02-11.xlsx")
